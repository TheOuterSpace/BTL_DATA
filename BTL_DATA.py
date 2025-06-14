import streamlit as st
import pandas as pd
import os
from PIL import Image
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Configuration
EXCEL_FILE = "shops_data.xlsx"
TEMP_IMAGE = "temp_img.jpg"
MAX_IMAGE_HEIGHT = 300  # pixels
IMAGE_COLUMN = 'C'  # Column for images
LAST_UPDATED_COLUMN = 'D'  # Column for timestamp
DATA_COLUMNS = ['Shop_ID', 'Region', 'last_updated']

def load_or_create_excel():
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if not hasattr(wb, 'max_image_width'):
            wb.max_image_width = 30  # Default width if not set
        return wb
    else:
        wb = load_workbook()
        ws = wb.active
        ws.append(DATA_COLUMNS)
        # Set initial column widths
        ws.column_dimensions['A'].width = 15  # Shop_ID
        ws.column_dimensions['B'].width = 20  # Region
        ws.column_dimensions[IMAGE_COLUMN].width = 30  # Images
        ws.column_dimensions[LAST_UPDATED_COLUMN].width = 20  # Timestamp
        wb.max_image_width = 30
        wb.save(EXCEL_FILE)
        return wb

def save_image_to_excel(shop_id, region, uploaded_file):
    try:
        wb = load_or_create_excel()
        ws = wb.active
        
        # Find existing row or create new
        row_idx = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0] == shop_id and row[1] == region:
                row_idx = idx
                break
        
        if not row_idx:
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value=shop_id)
            ws.cell(row=row_idx, column=2, value=region)
        
        # Process image
        img = Image.open(uploaded_file)
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        
        # Calculate dimensions maintaining aspect ratio
        if img.height > MAX_IMAGE_HEIGHT:
            ratio = MAX_IMAGE_HEIGHT / img.height
            new_height = MAX_IMAGE_HEIGHT
            new_width = int(img.width * ratio)
        else:
            new_width = img.width
            new_height = img.height
        
        img = img.resize((new_width, new_height))
        
        # Save temp image
        img.save(TEMP_IMAGE, "JPEG", quality=90)
        
        # Calculate required column width (7 pixels ≈ 1 Excel width unit)
        required_width = max(10, new_width / 7)
        
        # Update max width if this image is wider
        if required_width > wb.max_image_width:
            wb.max_image_width = required_width
            ws.column_dimensions[IMAGE_COLUMN].width = wb.max_image_width
        
        # Set row height (1.33 pixels ≈ 1 Excel height unit)
        row_height = max(15, new_height / 1.33)
        ws.row_dimensions[row_idx].height = row_height
        
        # Add image
        excel_img = ExcelImage(TEMP_IMAGE)
        ws.add_image(excel_img, f"{IMAGE_COLUMN}{row_idx}")
        
        # Add timestamp with proper alignment
        timestamp = datetime.now().strftime("%Y-%m-%d\n%H:%M:%S")
        timestamp_cell = ws.cell(row=row_idx, column=4, value=timestamp)
        timestamp_cell.alignment = Alignment(
            vertical='top',
            wrap_text=True
        )
        
        # Adjust timestamp column width to fit content
        ws.column_dimensions[LAST_UPDATED_COLUMN].width = max(
            20,  # Minimum width
            len("YYYY-MM-DD") + 2  # Enough for date + padding
        )
        
        # Center align other cells vertically
        for col in [1, 2]:  # Shop_ID and Region columns
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical='center')
        
        wb.save(EXCEL_FILE)
        return True
        
    except Exception as e:
        st.error(f"Error saving image: {str(e)}")
        return False
    finally:
        if os.path.exists(TEMP_IMAGE):
            os.remove(TEMP_IMAGE)

def main():
    st.title("Shop Photo Upload System")
    
    # Initialize session state for reset
    if 'reset_trigger' not in st.session_state:
        st.session_state.reset_trigger = False
    if 'selected_region' not in st.session_state:
        st.session_state.selected_region = None
    if 'selected_shop' not in st.session_state:
        st.session_state.selected_shop = None
    
    # Navigation
    app_mode = st.sidebar.radio("Navigation", ["Upload Photo", "View Data"])
    
    if app_mode == "Upload Photo":
        wb = load_or_create_excel()
        ws = wb.active
        
        # Get existing data for dropdowns
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row[:len(DATA_COLUMNS)] if row else DATA_COLUMNS)
        
        df = pd.DataFrame(data[1:], columns=DATA_COLUMNS[:len(data[0])])
        
        regions = sorted(df["Region"].unique()) if not df.empty else []
        shop_ids = sorted(df["Shop_ID"].unique()) if not df.empty else []
        
        if not regions or not shop_ids:
            st.warning("No shop data found. Please add shop data first.")
            return
        
        # Region selection
        selected_region = st.selectbox(
            "Select Region",
            regions,
            index=regions.index(st.session_state.selected_region) 
            if st.session_state.selected_region in regions else 0
        )
        
        # Shop selection filtered by region
        filtered_shops = df[df["Region"] == selected_region]["Shop_ID"].unique()
        selected_shop = st.selectbox(
            "Select Shop ID",
            filtered_shops,
            index=filtered_shops.tolist().index(st.session_state.selected_shop) 
            if st.session_state.selected_shop in filtered_shops else 0
        )
        
        # File uploader - key is controlled by reset_trigger
        uploader_key = "file_uploader_" + str(st.session_state.reset_trigger)
        uploaded_file = st.file_uploader(
            "Upload Shop Photo", 
            type=["jpg", "jpeg", "png"],
            key=uploader_key
        )
        
        if uploaded_file is not None:
            img = Image.open(uploaded_file)
            st.image(img, caption=f"Preview for Shop {selected_shop}")
            
            if st.button("Save Photo"):
                if save_image_to_excel(selected_shop, selected_region, uploaded_file):
                    st.success("Photo saved successfully!")
                    st.info(f"Image column width: {wb.max_image_width:.1f}")
                    
                    # Store current selections
                    st.session_state.selected_region = selected_region
                    st.session_state.selected_shop = selected_shop
                    
                    # Trigger UI reset by changing the uploader key
                    st.session_state.reset_trigger = not st.session_state.reset_trigger
                    st.rerun()
                else:
                    st.error("Failed to save photo")
    
    elif app_mode == "View Data":
        wb = load_or_create_excel()
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row[:len(DATA_COLUMNS)] if row else DATA_COLUMNS)
        
        df = pd.DataFrame(data[1:], columns=DATA_COLUMNS[:len(data[0])])
        st.dataframe(df)
        
        st.download_button(
            "Download Excel File",
            open(EXCEL_FILE, "rb").read(),
            EXCEL_FILE,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    main()
